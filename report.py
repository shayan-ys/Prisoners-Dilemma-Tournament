from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from datetime import datetime


def report_latex_table(tour):
    pop_reports_len = len(tour.pop_seed_memory)
    strategies_history = {}
    indices = []
    table_iteration_count = max(int(pop_reports_len / 4), 1)
    j = 0
    while j < pop_reports_len:
        if j + table_iteration_count >= pop_reports_len and j != pop_reports_len - 1:
            j = pop_reports_len - 1

        for strategy, mem_count in tour.pop_seed_memory[j].items():
            try:
                strategies_history[strategy.name].append(mem_count)
            except KeyError:
                strategies_history[strategy.name] = [mem_count]
        indices.append(j)

        j += table_iteration_count

    print('\\begin{tabular}{' + '|c|' * (len(indices) + 1) + '}')
    print('\\hline')
    print('Population & ' + ' & '.join(map(str, indices)) + '\\\\')
    print('\\hline')
    for strategy, pop_history in strategies_history.items():
        print(strategy + ' & ' + ' & '.join(map(str, pop_history)) + '\\\\')
        print('\\hline')
    print('\\end{tabular}')


def report_to_spreadsheet(tour, **kwargs):
    wb = Workbook()
    ws = wb.active
    cs = wb.create_chartsheet()

    chart = LineChart()

    headers = []
    header_seed_keys = []

    seeds_history = tour.pop_seed_memory

    for key, value in seeds_history[0].items():
        if not value:
            continue
        headers.append(key.name)
        header_seed_keys.append(key)

    ws.append(headers)

    for row_index, seed in enumerate(seeds_history):
        row = []

        for key in header_seed_keys:
            try:
                row.append(seed[key])
            except KeyError:
                pass

        ws.append(row)

    data = Reference(ws, min_col=1, max_col=len(headers), min_row=1, max_row=len(seeds_history))
    chart.add_data(data, titles_from_data=True)

    cs.add_chart(chart)

    row_index = 2
    for label, value in kwargs.items():
        ws.cell(row=row_index, column=len(headers) + 2, value=label)
        ws.cell(row=row_index, column=len(headers) + 3, value=value)
        row_index += 1

    dnow = datetime.now()
    report_name = str(dnow.day) + '_' + str(dnow.hour) + '_' + str(dnow.minute) + '_' + str(dnow.second)

    wb.save('reports/' + report_name + '.xlsx')
